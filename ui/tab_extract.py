# ui/tab_extract.py
"""
TabExtract  (PDF → OCR → Excel 탭, Drag&Drop 지원판)
───────────────────────────────────────────────────
• ROI 세트 선택
• PDF 파일(여러 개) 선택  ─ 버튼·Drag&Drop 모두 지원
• 엑셀 파일 + 시트 선택  ─ 버튼·Drag&Drop 모두 지원
• 매핑 테이블에서 ‘엑셀 열’ 직접 입력
• 전역 제외 문자열 적용 후 원본 보존 + 새 파일로 저장
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFileDialog, QTextEdit, QTableWidget, QTableWidgetItem,
    QMessageBox
)

import fitz  # PyMuPDF
from openpyxl.utils import column_index_from_string

from core.models import ROISet
from core.roi_manager import ROIManager
from core.exclusion_manager import ExclusionManager
from core.ocr_engine import OCREngine
from core.excel_writer import ExcelWriter


class TabExtract(QWidget):
    def __init__(
        self,
        roi_mgr: ROIManager,
        ex_mgr: ExclusionManager,
        parent: QWidget | None = None
    ):
        super().__init__(parent)
        self.roi_mgr = roi_mgr
        self.ex_mgr  = ex_mgr
        self.ocr = OCREngine(dpi=600, lang="kor+eng")

        # ─ 위젯 ─
        self.cmb_set   = QComboBox()
        self.btn_pdf   = QPushButton("PDF 선택")
        self.lbl_pdf   = QLabel("선택 없음")
        self.btn_excel = QPushButton("엑셀 선택")
        self.cmb_sheet = QComboBox()
        self.table_map = QTableWidget()
        self.btn_run   = QPushButton("변환 실행")
        self.log       = QTextEdit()
        self.log.setReadOnly(True)

        # ─ 레이아웃 ─
        h1 = QHBoxLayout(); h1.addWidget(QLabel("ROI 세트:")); h1.addWidget(self.cmb_set); h1.addStretch()
        h2 = QHBoxLayout(); h2.addWidget(self.btn_pdf); h2.addWidget(self.lbl_pdf, 1)
        h3 = QHBoxLayout(); h3.addWidget(self.btn_excel); h3.addWidget(QLabel("시트:")); h3.addWidget(self.cmb_sheet, 1)
        h4 = QHBoxLayout(); h4.addStretch(); h4.addWidget(self.btn_run)

        v = QVBoxLayout(self)
        v.addLayout(h1)
        v.addLayout(h2)
        v.addLayout(h3)
        v.addWidget(QLabel("ROI ↔ 엑셀 열 매핑 (A, B … 또는 헤더명):"))
        v.addWidget(self.table_map)
        v.addLayout(h4)
        v.addWidget(self.log)

        # 테이블 설정
        self.table_map.setColumnCount(2)
        self.table_map.setHorizontalHeaderLabels(["ROI 이름", "엑셀 열"])
        self.table_map.setColumnWidth(0, 200)

        # 초기화: 세트 목록 로드 및 매핑 적용
        names = self.roi_mgr.list_sets()
        self.cmb_set.addItems(names)
        self.cmb_set.currentTextChanged.connect(self.populate_mapping)
        if names:
            self.populate_mapping(self.cmb_set.currentText())

        # 시그널 연결
        self.btn_pdf.clicked.connect(self.select_pdf)
        self.btn_excel.clicked.connect(self.select_excel)
        self.btn_run.clicked.connect(self.run_conversion)

        # Drag&Drop 설정
        self.setAcceptDrops(True)
        self.pdf_paths: List[Path] = []
        self.excel_path: Path | None = None

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            for u in e.mimeData().urls():
                f = u.toLocalFile().lower()
                if f.endswith((".pdf", ".xlsx", ".xls", ".xlsm")):
                    e.acceptProposedAction()
                    return
        e.ignore()

    def dropEvent(self, e):
        for url in e.mimeData().urls():
            f = Path(url.toLocalFile())
            if f.suffix.lower() == ".pdf":
                self.pdf_paths = [f]
                self.lbl_pdf.setText(f.name)
            elif f.suffix.lower() in (".xlsx", ".xls", ".xlsm"):
                self.excel_path = f
                self.load_sheets()
        if self.excel_path and not self.cmb_sheet.count():
            self.load_sheets()

    def populate_mapping(self, set_name: str) -> None:
        rs: ROISet | None = self.roi_mgr.get_set(set_name)
        if not rs:
            self.table_map.setRowCount(0)
            return
        self.table_map.setRowCount(len(rs.rois))
        for row, r in enumerate(rs.rois):
            self.table_map.setItem(row, 0, QTableWidgetItem(r.name))
            editable = QTableWidgetItem("")
            editable.setTextAlignment(Qt.AlignCenter)
            editable.setFlags(editable.flags() | Qt.ItemIsEditable)
            self.table_map.setItem(row, 1, editable)

    def select_pdf(self) -> None:
        files, _ = QFileDialog.getOpenFileNames(self, "PDF 선택", "", "PDF Files (*.pdf)")
        if files:
            self.pdf_paths = [Path(f) for f in files]
            self.lbl_pdf.setText(f"{len(files)}개 PDF 선택")

    def select_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "엑셀 선택", "", "Excel Files (*.xlsx *.xls *.xlsm)")
        if path:
            self.excel_path = Path(path)
            self.load_sheets()

    def load_sheets(self) -> None:
        if not self.excel_path:
            return
        import openpyxl, pandas as pd
        if self.excel_path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            sheets = wb.sheetnames
        else:
            xls = pd.ExcelFile(self.excel_path)
            sheets = xls.sheet_names
        self.cmb_sheet.clear()
        self.cmb_sheet.addItems(sheets)

    def run_conversion(self) -> None:
        # 필수 선택 확인
        set_name = self.cmb_set.currentText()
        rs = self.roi_mgr.get_set(set_name)
        if not rs:
            QMessageBox.warning(self, "실행 불가", "ROI 세트 선택 필요")
            return
        if not self.pdf_paths:
            QMessageBox.warning(self, "실행 불가", "PDF 선택 필요")
            return
        if not self.excel_path:
            QMessageBox.warning(self, "실행 불가", "엑셀 선택 필요")
            return

        # 매핑 dict 생성 및 누락 체크
        mapping: Dict[str, str] = {}
        for row in range(self.table_map.rowCount()):
            roi_name = self.table_map.item(row, 0).text()
            col_text = self.table_map.item(row, 1).text().strip().upper()
            if not col_text:
                QMessageBox.warning(self, "누락", "모든 ROI의 엑셀 열을 입력하세요.")
                return
            mapping[roi_name] = col_text

        # 엑셀 워크북 로드
        writer = ExcelWriter(self.excel_path)
        wb = writer.wb
        sheet_name = self.cmb_sheet.currentText()
        ws = wb[sheet_name]
        ex_set = set(self.ex_mgr.list_all())

        # 시작 행 결정: 매핑된 열 중 가장 마지막 데이터가 있는 행 +1
        col_idxs = [column_index_from_string(col) for col in mapping.values()]
        last_rows: List[int] = []
        for idx in col_idxs:
            found = 0
            for r in range(ws.max_row, 0, -1):
                if ws.cell(row=r, column=idx).value not in (None, ""):
                    found = r
                    break
            last_rows.append(found)
        new_row = max(last_rows, default=0) + 1

        # PDF 파일 및 페이지별로 OCR → 엑셀 삽입
        for pdf in self.pdf_paths:
            doc = fitz.open(pdf)
            for page_num in range(doc.page_count):
                # 테이블 ROI 단독 처리
                table_rois = [r for r in rs.rois if r.field_type == 'table']
                if table_rois and len(rs.rois) == len(table_rois):
                    r = table_rois[0]
                    txt = self.ocr.extract_roi(pdf, page_num, (r.x, r.y, r.w, r.h), r.tolerance)
                    lines = txt.strip().splitlines()
                    for line in lines:
                        col_idx = column_index_from_string(mapping[r.name])
                        ws.cell(row=new_row, column=col_idx, value=line)
                        self.log.append(f"{pdf.name} 페이지 {page_num+1} 테이블 행 {new_row}")
                        new_row += 1
                    continue

                # 혼합 또는 단일 ROI 처리: 한 페이지당 한 행 삽입
                data: Dict[str, str] = {}
                for roi in rs.rois:
                    txt = self.ocr.extract_roi(pdf, page_num, (roi.x, roi.y, roi.w, roi.h), roi.tolerance)
                    cleaned = txt.strip()
                    data[roi.name] = "" if any(ex in cleaned for ex in ex_set) else cleaned

                for roi_name, col_letter in mapping.items():
                    col_idx = column_index_from_string(col_letter)
                    ws.cell(row=new_row, column=col_idx, value=data[roi_name])
                self.log.append(f"{pdf.name} 페이지 {page_num+1} 처리 완료 (행 {new_row})")
                new_row += 1
            doc.close()

        # 새 파일로 저장, 원본 보존
        saved = writer.save_as_new()
        self.log.append(f"→ 저장 완료: {saved.name}")
        QMessageBox.information(self, "완료", f"파일이 저장되었습니다:\n{saved}")

    def refresh_sets(self) -> None:
        """외부 호출용: 좌표 세트 목록 갱신"""
        names = self.roi_mgr.list_sets()
        self.cmb_set.blockSignals(True)
        self.cmb_set.clear()
        self.cmb_set.addItems(names)
        self.cmb_set.blockSignals(False)
        if names:
            self.populate_mapping(self.cmb_set.currentText())
