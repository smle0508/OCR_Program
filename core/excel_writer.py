# core/excel_writer.py
"""
ExcelWriter
───────────
• 항상 새로운 엑셀(.xlsx) 워크북을 생성하여 데이터 기록
• 단일 셀, 튜플 좌표, 표(table) 형식 동시 지원
• 저장 시 지정한 경로에 .xlsx로 저장
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from typing import Any, Dict, List, Union


class ExcelWriter:
    """
    새로운 워크북을 생성하여 데이터를 기록하고 저장하는 유틸리티 클래스

    Attributes:
        output_path (Path): 저장할 파일 경로
        wb (Workbook): openpyxl 워크북 객체
        ws (Worksheet): 현재 활성 시트 객체
    """

    def __init__(
        self,
        output_path: Union[str, Path],
        sheet_name: str = "Sheet1"
    ):
        # 출력 경로 설정
        self.output_path = Path(output_path)
        # 새 워크북 생성 및 시트명 설정
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = sheet_name
        self.ws = ws

    def write_values(self, values: Dict[Union[str, tuple[int, int]], Any]) -> None:
        """
        단일 셀 또는 (row, col) 좌표 튜플에 값을 기록합니다.

        Args:
            values (Dict[Union[str, tuple[int, int]], Any]):
                키로 엑셀 셀 주소 문자열(A1 등) 또는 (row, col) 1-based 튜플을 허용
                값은 셀에 기록할 데이터
        """
        for cell, val in values.items():
            if isinstance(cell, tuple) and len(cell) == 2 and all(isinstance(i, int) for i in cell):
                # (row, column) 튜플 주소 처리
                row, col = cell
                self.ws.cell(row=row, column=col, value=val)
            else:
                # 문자열 주소(A1 등)
                self.ws[cell] = val

    def write_table(self, start_cell: str, table_data: List[List[Any]]) -> None:
        """
        표 형식 데이터를 지정한 시작 셀에서부터 기록합니다.

        Args:
            start_cell (str): 시작 셀 주소 (예: "A5")
            table_data (List[List[Any]]): 2차원 리스트 형태의 테이블 데이터
        """
        # 열 문자와 행 번호 분리
        col_letters = ''.join(filter(str.isalpha, start_cell))
        row_number = int(''.join(filter(str.isdigit, start_cell)))
        start_col = self._col_letter_to_index(col_letters)

        # 테이블 데이터 기록
        for r_idx, row in enumerate(table_data):
            for c_idx, val in enumerate(row):
                self.ws.cell(
                    row=row_number + r_idx,
                    column=start_col + c_idx,
                    value=val
                )

    def save(self) -> None:
        """
        워크북을 지정된 경로에 .xlsx로 저장합니다.
        """
        # 확장자가 .xlsx가 아니면 자동 변경
        if self.output_path.suffix.lower() != ".xlsx":
            self.output_path = self.output_path.with_suffix(".xlsx")
        self.wb.save(self.output_path)

    def _col_letter_to_index(self, letters: str) -> int:
        """
        엑셀 열 문자를 숫자 인덱스(1-based)로 변환합니다.

        예: A->1, B->2, ..., Z->26, AA->27
        """
        letters = letters.upper()
        index = 0
        for ch in letters:
            index = index * 26 + (ord(ch) - ord('A') + 1)
        return index
